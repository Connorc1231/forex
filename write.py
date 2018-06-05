from crontab import CronTab
import datetime
from pathlib import Path

import xlsxwriter
import xlrd
from xlutils.copy import copy
import openpyxl

import main as indicators



def main():
  result = indicators.main()
  my_file = Path('Documents/forex/indicators.xlsx')

  if not my_file.is_file():
    createSpreadsheet(result)

  appendToSpreadsheet(result)

def createSpreadsheet(indicators):
  wb = openpyxl.Workbook()
  ws = wb.active
  ws['A1'] = 'Time'
  ws['B1'] = 'SMA'
  ws['C1'] = 'EMA'
  ws['D1'] = 'RSI'
  ws['E1'] = 'Stoch'
  ws['F1'] = 'Bbands Upper'
  ws['G1'] = 'Bbands Mid'
  ws['H1'] = 'Bbands Lower'

  ws.column_dimensions["A"].width = 20
  ws.column_dimensions["F"].width = 12
  ws.column_dimensions["G"].width = 12
  ws.column_dimensions["H"].width = 12

  wb.save(filename = 'Documents/forex/indicators.xlsx')
  print('Created indicators.xlsx')

def appendToSpreadsheet(indicators):
  wb = openpyxl.load_workbook('Documents/forex/indicators.xlsx')
  ws = wb.active

  # New data to write:
  inputList = [
    datetime.datetime.now(),
    indicators['SMA'],
    indicators['EMA'],
    indicators['RSI'],
    indicators['Stoch'],
    indicators['BBands']['upper'],
    indicators['BBands']['mid'],
    indicators['BBands']['lower']
  ]

  ws.append(inputList)    
  wb.save(filename = 'Documents/forex/indicators.xlsx')
  print('Appended to indicators.xlsx')

if __name__ == "__main__":
  main()